# 生成「行政部驾驶舱」WPS在线表格模板 (v1.1.0)
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference

N = 500  # 预置行数

wb = Workbook()

# ---------- 表1：任务总表 ----------
ws = wb.active
ws.title = "任务总表"

headers = ["任务名称","所属大类","子项备注","负责人","协同人","状态","进度%",
           "开始日期","截止日期","优先级","堵点说明","是否逾期","更新日期"]
ws.append(headers)

# 表头样式
hdr_fill = PatternFill("solid", fgColor="2B67FF")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
for c in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# 各大类 / 状态 / 优先级 选项
big_cats = "会务接待,车辆管理,后勤管理,物资采购,行政办公,人力资源,其他"
statuses = "未开始,进行中,已完成,已暂停,已逾期"
prios = "高,中,低"

def add_dv(col_letter, formula):
    dv = DataValidation(type="list", formula1='"' + formula + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{N+1}")

add_dv("B", big_cats)   # 所属大类
add_dv("F", statuses)   # 状态
add_dv("J", prios)      # 优先级

# 是否逾期公式 + 进度%格式 + 日期格式
for r in range(2, N+2):
    ws.cell(row=r, column=12).value = (  # L 是否逾期
        f'=IF(AND(F{r}<>"已完成",I{r}<>"",TODAY()>I{r}),"逾期","正常")')
    ws.cell(row=r, column=7).number_format = '0"%"'      # G 进度%
    ws.cell(row=r, column=8).number_format = "yyyy-mm-dd" # H 开始
    ws.cell(row=r, column=9).number_format = "yyyy-mm-dd" # I 截止

# 条件格式
red = PatternFill("solid", fgColor="FFC7CE")
green = PatternFill("solid", fgColor="C6EFCE")
yellow = PatternFill("solid", fgColor="FFEB9C")
# 逾期整行标红
ws.conditional_formatting.add(f"A2:M{N+1}",
    FormulaRule(formula=['$L2="逾期"'], fill=red))
# 已完成 绿 / 进行中 黄（仅 F 列）
ws.conditional_formatting.add(f"F2:F{N+1}",
    FormulaRule(formula=['$F2="已完成"'], fill=green))
ws.conditional_formatting.add(f"F2:F{N+1}",
    FormulaRule(formula=['$F2="进行中"'], fill=yellow))
# 进度% 数据条
ws.conditional_formatting.add(f"G2:G{N+1}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100,
                color="638EC6", showValue=True))

# 列宽
widths = [22,12,20,10,14,10,9,13,13,9,22,10,13]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w
ws.freeze_panes = "A2"

# ---------- 表2：仪表盘 ----------
dash = wb.create_sheet("仪表盘")
dash["A1"] = "行政部驾驶舱 · 仪表盘"
dash["A1"].font = Font(bold=True, size=14, color="2B67FF")

# 统计卡
dash["A3"]="总任务数"; dash["B3"]=f'=COUNTA(任务总表!A2:A{N+1})'
dash["A4"]="已完成";   dash["B4"]=f'=COUNTIF(任务总表!F2:F{N+1},"已完成")'
dash["A5"]="进行中";   dash["B5"]=f'=COUNTIF(任务总表!F2:F{N+1},"进行中")'
dash["A6"]="已暂停";   dash["B6"]=f'=COUNTIF(任务总表!F2:F{N+1},"已暂停")'
dash["A7"]="未开始";   dash["B7"]=f'=COUNTIF(任务总表!F2:F{N+1},"未开始")'
dash["A8"]="逾期";     dash["B8"]=f'=COUNTIF(任务总表!L2:L{N+1},"逾期")'
dash["A9"]="完成率";   dash["B9"]=f'=IF(B3>0,B4/B3,0)'; dash["B9"].number_format='0%'
for r in range(3,10):
    dash.cell(row=r, column=1).font = Font(bold=True)
    dash.cell(row=r, column=2).font = Font(size=12)

# 状态分布（饼图数据）
dash["D3"]="状态"; dash["E3"]="数量"
dash["D4"]="未开始"; dash["E4"]="=B7"
dash["D5"]="进行中"; dash["E5"]="=B5"
dash["D6"]="已完成"; dash["E6"]="=B4"
dash["D7"]="已暂停"; dash["E7"]="=B6"
dash["D8"]="逾期";   dash["E8"]="=B8"

# 各大类分布（柱状图数据）
cats = ["会务接待","车辆管理","后勤管理","物资采购","行政办公","人力资源","其他"]
dash["G3"]="大类"; dash["H3"]="数量"
for i, c in enumerate(cats):
    dash[f"G{4+i}"]=c
    dash[f"H{4+i}"]=f'=COUNTIF(任务总表!B2:B{N+1},G{4+i})'

# 饼图：状态分布
pie = PieChart(); pie.title="任务状态分布"
data = Reference(dash, min_col=5, min_row=3, max_row=8)
cats_ref = Reference(dash, min_col=4, min_row=4, max_row=8)
pie.add_data(data, titles_from_data=True); pie.set_categories(cats_ref)
pie.height=7; pie.width=11
dash.add_chart(pie, "D10")

# 柱状图：各大类
bar = BarChart(); bar.title="各大类任务数"; bar.type="col"; bar.legend=None
bdata = Reference(dash, min_col=8, min_row=3, max_row=10)
bcats = Reference(dash, min_col=7, min_row=4, max_row=10)
bar.add_data(bdata, titles_from_data=True); bar.set_categories(bcats)
bar.height=7; bar.width=11
dash.add_chart(bar, "G10")

for col,w in zip("ABCDEFGH",[10,10,10,10,10,10,12,10]):
    dash.column_dimensions[col].width = w

out = r"D:\微云文件\知识库\落地的做法\行政部的驾驶舱\行政部驾驶舱模板_v1.1.0.xlsx"
wb.save(out)
print("已生成:", out)
